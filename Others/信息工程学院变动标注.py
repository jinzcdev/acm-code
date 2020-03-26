# -*- coding: utf-8 -*-
from time import strftime, localtime
from tkinter import filedialog, Tk, Label, Entry, Button, StringVar
from tkinter import messagebox

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.fonts import Font

# 定义文件类型
file_type = [('Excel文件类型', '.xlsx')]


# 需要检查的列号
check_col_numbers = [9, 10, 11, 12, 13, 14, 15, 16, 20, 21, 22,
                     23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36]


# 打开文件选择框, 选择对比文件路径
def select_opened_file_path(sv):
    sv.set(filedialog.askopenfilename(
        title='选择文件路径',
        filetypes=file_type,
        defaultextension='.xlsx'))


# 打开文件保存框, 选择导出文件路径
def select_saved_file_path(sv):
    current_date = strftime("%Y_%m_%d", localtime())
    sv.set(filedialog.asksaveasfilename(
        title='选择导出文件路径',
        filetypes=file_type,
        defaultextension='.xlsx',
        initialfile='信息工程学院' + current_date + '学生信息填报.xlsx'))


# 对比处理函数, 导出结果
def solve():
    cnt = 0
    try:
        # 存在空路径退出函数
        if file_today.get() == '' or file_yesterday.get() == '' or file_output.get() == '':
            messagebox.showinfo("错误信息", "文件路径有误, 请重新选择!")
            return
        # 今日与昨日工作簿
        wb_today = load_workbook(filename=file_today.get())
        wb_yesterday = load_workbook(filename=file_yesterday.get())

        # 今日与昨日表
        sheet_today = wb_today.worksheets[0]
        sheet_yesterday = wb_yesterday.worksheets[0]

        # 行号起始与结束编号
        index_start, index_end = 5, sheet_today.max_row

        for row_num in range(index_start, index_end + 1):
            is_changed = False
            list_diff = []
            for col_num in check_col_numbers:
                if sheet_today.cell(row=row_num, column=col_num).value != sheet_yesterday.cell(row=row_num, column=col_num).value:
                    # 列表记录不同单元格列号
                    list_diff.append(col_num)
                    is_changed = True
            # 不是完全相同的, 将该行填充为红色, 再将不同单元格填充为黄色
            if is_changed:
                cnt = cnt + 1
                font = sheet_today.cell(row=row_num, column=col_num).font.copy(color='00000000')
                for col_num in range(1, 37 + 1):
                    sheet_today.cell(row=row_num, column=col_num).fill = PatternFill('solid', fgColor='FF0000')
                    sheet_today.cell(row=row_num, column=col_num).font = font
                for col_num in list_diff:
                    sheet_today.cell(row=row_num, column=col_num).fill = PatternFill('solid', fgColor='FFFF00')
        wb_today.save(file_output.get())
    except:
        messagebox.showinfo("错误信息", "文件路径有误, 请重新选择!")
    else:
        wb_today.close()
        wb_yesterday.close()
        messagebox.showinfo("提示信息", '已导出至储存路径, 共 %d 人信息变动' % cnt)


# tkinter库创建简单图形界面
root = Tk()
root.title('信息工程学院-学生信息改动标注 powered by 金')
root.resizable(0, 0)
file_today, file_yesterday, file_output = StringVar(), StringVar(), StringVar()
Label(root, text='注: 仅支持.xlsx格式的Excel表格, .xls另存为至.xlsx做格式转换').grid(row=0, column=1)
Label(root, text='昨日文件路径:').grid(row=1, column=0)
Entry(root, textvariable=file_yesterday).grid(row=1, column=1, sticky='we')
Button(root, text='路径选择', command=lambda: select_opened_file_path(file_yesterday)).grid(row=1, column=2)
Label(root, text='今日文件路径:').grid(row=2, column=0)
Entry(root, textvariable=file_today).grid(row=2, column=1, sticky='we')
Button(root, text='路径选择', command=lambda: select_opened_file_path(file_today)).grid(row=2, column=2)
Label(root, text='保存路径:').grid(row=3, column=0)
Entry(root, textvariable=file_output).grid(row=3, column=1, sticky='we')
Button(root, text='路径选择', command=lambda: select_saved_file_path(file_output)).grid(row=3, column=2)

Button(root, text='点我导出 (非程序崩溃, Python程序需要1分钟处理数据)', command=solve).grid(row=4, column=1)
root.mainloop()